#!/usr/bin/env python3
"""03 — Validate the reviewed workbook.

    python scripts/03_validate_sheet.py
    python scripts/03_validate_sheet.py --approve --by "Your Name"

Catches contradictions before anything touches an org. Runs in seconds and
saves the expensive failure: a half-applied deploy.

Exit codes:  0 clean (warnings allowed)   1 failures found   2 cannot read

CHECKPOINT: set a field to Y under an object set to N. Confirm this exits
non-zero naming the tab and row.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from lib import hashing, mapping, workbook as wbmod

ROOT = Path(__file__).resolve().parent.parent
WB = ROOT / "mapping" / "workbook.xlsx"


class Issue:
    def __init__(self, level, tab, row, problem, fix):
        self.level = level      # FAIL or WARN
        self.tab = tab
        self.row = row
        self.problem = problem
        self.fix = fix

    def __str__(self):
        return f"  {self.tab}!row {self.row}\n      {self.problem}\n      FIX: {self.fix}"


class Validator:
    def __init__(self, tabs, config):
        self.tabs = tabs
        self.config = config
        self.naming = config["naming"]
        self.issues: list[Issue] = []

    def fail(self, row, problem, fix):
        self.issues.append(Issue("FAIL", row.get("_tab", "?"), row.get("_row", "?"),
                                 problem, fix))

    def warn(self, row, problem, fix):
        self.issues.append(Issue("WARN", row.get("_tab", "?"), row.get("_row", "?"),
                                 problem, fix))

    @property
    def failures(self):
        return [i for i in self.issues if i.level == "FAIL"]

    @property
    def warnings(self):
        return [i for i in self.issues if i.level == "WARN"]

    # -- helpers -----------------------------------------------------------

    @staticmethod
    def deploy(row) -> str:
        return str(row.get("Deploy", "")).strip().upper()

    @staticmethod
    def final(row, column, fallback_column=None):
        value = row.get(column)
        if value in (None, "") and fallback_column:
            value = row.get(fallback_column)
        return str(value or "").strip()

    # -- checks ------------------------------------------------------------

    def check_holds(self):
        """Nothing deploys while anyone is undecided."""
        for row in wbmod.holds(self.tabs):
            label = (row.get("HS Property") or row.get("HS Association")
                     or row.get("HS Stage") or row.get("Rule_Name") or "?")
            self.fail(row, f"{label!r} is still at HOLD.",
                      "Set Deploy to Y or N.")

    def check_object_consistency(self):
        """A field marked Y under an object marked N is an orphan.

        The tab's object-level decision is the FIELDS row for the external-id
        field where present, otherwise: an object is 'off' when every one of
        its field rows is N.
        """
        for tab_name, tab in self.tabs.items():
            fields = tab.sections.get("FIELDS")
            if not fields or not fields.rows:
                continue

            # The object is "off" when not one of its fields is deploying.
            # Anything else on the tab marked Y is then an orphan.
            object_off = not any(self.deploy(r) == "Y" for r in fields.rows)
            if not object_off:
                continue

            for section_name in ("PICKLISTS", "RELATIONSHIPS", "RECORD_TYPES",
                                 "VALIDATION_RULES"):
                section = tab.sections.get(section_name)
                if not section:
                    continue
                for row in section.rows:
                    if self.deploy(row) == "Y":
                        self.fail(row,
                                  f"{section_name} row marked Y, but every field on "
                                  f"{tab_name} is N — the object is not being deployed.",
                                  f"Set this to N, or set at least one {tab_name} "
                                  "field to Y.")

    def check_api_names(self):
        limits = self.naming["limits"]
        reserved = {str(w).lower() for w in self.naming.get("reserved_words", [])}
        seen: dict[tuple[str, str], str] = {}

        for row in wbmod.rows_where(self.tabs, section="FIELDS", deploy="Y"):
            api = self.final(row, "Final_API", "Proposed_API")
            hs = row.get("HS Property", "?")
            tab = row.get("_tab", "?")

            if not api:
                self.fail(row, f"{hs}: no API name.", "Set Final_API.")
                continue

            # Standard fields carry no __c and are exempt from custom rules.
            if not api.endswith("__c"):
                proposed_type = str(row.get("Proposed_Type", "")).strip()
                if proposed_type == "Standard":
                    continue
                self.fail(row, f"{api!r} does not end with __c.",
                          f"Change Final_API to {api}__c, or map it to a standard field.")
                continue

            if len(api) > limits["field_api_name_max"]:
                self.fail(row,
                          f"{api!r} is {len(api)} chars, max {limits['field_api_name_max']}.",
                          "Shorten Final_API.")

            base = api[:-3]
            if base.lower() in reserved:
                self.fail(row, f"{base!r} is a reserved Salesforce word.",
                          f"Rename, e.g. {base}_HS__c.")
            if "__" in base:
                self.fail(row, f"{api!r} contains a double underscore before __c.",
                          "Remove the doubled underscore.")
            if base.startswith("_") or base.endswith("_"):
                self.fail(row, f"{api!r} has a leading or trailing underscore.",
                          "Remove it.")
            if base and not base[0].isalpha():
                self.fail(row, f"{api!r} does not start with a letter.",
                          "Prefix with a letter, e.g. X.")

            key = (tab, api.lower())
            if key in seen:
                self.fail(row,
                          f"{api!r} duplicates the API name already used by "
                          f"{seen[key]!r} on this object.",
                          "Give one of them a distinct Final_API.")
            else:
                seen[key] = hs

    def check_types(self):
        valid = set(self.config["types"]["sf_field_types"]) | {"Standard"}

        for row in wbmod.rows_where(self.tabs, section="FIELDS", deploy="Y"):
            sf_type = self.final(row, "Final_Type", "Proposed_Type")
            hs = row.get("HS Property", "?")

            if sf_type and sf_type not in valid:
                self.fail(row, f"{hs}: {sf_type!r} is not a valid Salesforce field type.",
                          f"Pick one of: {', '.join(sorted(valid))}.")
                continue

            length = self.final(row, "Final_Len", "Proposed_Len")
            if sf_type == "Text" and length.isdigit() and int(length) > 255:
                self.warn(row, f"{hs}: Text with length {length} exceeds the 255 limit.",
                          "Change Final_Type to LongTextArea.")

            if sf_type == "MasterDetail":
                self.warn(row,
                          f"{hs}: master-detail cannot be added to an object that "
                          "already has records, and cannot be undone.",
                          "Confirm the target object is empty, or use Lookup.")

    def check_picklists(self):
        limit = self.naming["limits"]["picklist_value_max"]
        seen: dict[tuple[str, str], set] = {}

        for row in wbmod.rows_where(self.tabs, section="PICKLISTS", deploy="Y"):
            value = self.final(row, "Final_Value", "Proposed_Value")
            prop = row.get("HS Property", "?")
            key = (row.get("_tab", "?"), prop)

            if not value:
                self.fail(row, f"{prop}: empty picklist value.", "Set Final_Value.")
                continue
            if len(value) > limit:
                self.fail(row, f"{prop}: value is {len(value)} chars, max {limit}.",
                          "Shorten Final_Value.")
            bucket = seen.setdefault(key, set())
            if value.lower() in bucket:
                self.fail(row, f"{prop}: duplicate picklist value {value!r}.",
                          "Remove or rename the duplicate.")
            bucket.add(value.lower())

    def check_relationships(self):
        """A lookup whose target is neither deploying nor already in the org."""
        deploying_objects = set()
        for tab_name, tab in self.tabs.items():
            fields = tab.sections.get("FIELDS")
            if fields and any(self.deploy(r) == "Y" for r in fields.rows):
                deploying_objects.add(tab_name.lower())

        standard = {"account", "contact", "opportunity", "case", "lead", "user",
                    "product2", "opportunitylineitem", "campaign", "task", "event"}

        for row in wbmod.rows_where(self.tabs, section="RELATIONSHIPS", deploy="Y"):
            target = self.final(row, "Final_Target", "Proposed_Target")
            if not target:
                self.fail(row, "Lookup has no target object.", "Set Final_Target.")
                continue

            key = target.lower()
            tab_key = target.replace("__c", "").lower()
            known = (key in standard
                     or key in deploying_objects
                     or tab_key in {t.replace("__c", "").lower() for t in deploying_objects}
                     or any(t.lower() == key for t in self.tabs))
            if not known:
                self.fail(row,
                          f"Lookup targets {target!r}, which is neither being deployed "
                          "nor a standard object.",
                          f"Deploy {target} first, or set this relationship to N.")

            if self.final(row, "Final_Type", "Proposed_Type") == "MasterDetail":
                self.fail(row,
                          f"Master-detail to {target}: impossible if the child object "
                          "already has records, and irreversible.",
                          "Use Lookup unless you are certain the object is empty.")

    def check_validation_rules(self):
        deployed_fields = {
            (r.get("_tab"), self.final(r, "Final_API", "Proposed_API").lower())
            for r in wbmod.rows_where(self.tabs, section="FIELDS", deploy="Y")
        }
        skipped = {
            (r.get("_tab"), self.final(r, "Final_API", "Proposed_API").lower())
            for r in wbmod.rows_where(self.tabs, section="FIELDS", deploy="N")
        }

        for row in wbmod.rows_where(self.tabs, section="VALIDATION_RULES", deploy="Y"):
            name = row.get("Rule_Name", "?")
            formula = self.final(row, "Final_Formula", "Proposed_Formula")
            tab = row.get("_tab")

            if not name:
                self.fail(row, "Validation rule marked Y with no name.",
                          "Set Rule_Name, or set Deploy to N.")
            if not formula:
                self.fail(row, f"{name}: no formula.",
                          "Set Final_Formula, or set Deploy to N.")
            if not self.final(row, "Error_Message_Text"):
                self.fail(row, f"{name}: no error message.",
                          "Set Error_Message_Text — Salesforce requires one.")

            for tab_key, field_api in skipped:
                if tab_key != tab or not field_api:
                    continue
                bare = field_api.replace("__c", "")
                if bare and bare in formula.lower() and (tab, field_api) not in deployed_fields:
                    self.fail(row,
                              f"{name}: references {field_api}, which is marked N.",
                              f"Set {field_api} to Y, or set this rule to N.")

            if str(row.get("Active", "")).strip().upper() in ("TRUE", "Y", "YES"):
                self.warn(row,
                          f"{name}: Active=TRUE. Rules written for future data entry "
                          "reject legitimate historical records.",
                          "Set Active to FALSE, and activate after the data load.")

    def check_required_fields(self):
        for row in wbmod.rows_where(self.tabs, section="FIELDS", deploy="Y"):
            notes = str(row.get("Notes", "")).lower()
            if "required" not in notes:
                continue
            fill = str(row.get("Fill %", "")).replace("%", "").strip()
            try:
                value = float(fill)
            except ValueError:
                continue
            if value < 100:
                self.warn(row,
                          f"{row.get('HS Property')}: marked required in Notes but only "
                          f"{value:.1f}% filled in HubSpot — the load WILL reject the gaps.",
                          "Leave it optional, or backfill in HubSpot first.")

    def check_owner_mapping(self, users):
        """hubspot_owner_id mapped but owners unmapped → ownership silently
        defaults to whoever runs the load."""
        owner_rows = [r for r in wbmod.rows_where(self.tabs, section="FIELDS", deploy="Y")
                      if str(r.get("HS Property", "")).lower() == "hubspot_owner_id"]
        if not owner_rows:
            return

        # "Archived" is the _Users column header — archived owners are
        # legitimately left unmapped, so only active ones can fail this.
        unmapped = [u for u in users
                    if str(u.get("Archived", "")).strip().upper() != "Y"
                    and not str(u.get("SF_Username", "")).strip()]
        if unmapped:
            row = owner_rows[0]
            names = ", ".join(f"{u.get('HS Email') or u.get('HS Name')}" for u in unmapped[:5])
            more = f" (+{len(unmapped) - 5} more)" if len(unmapped) > 5 else ""
            self.fail(row,
                      f"Owner field is marked Y but {len(unmapped)} active HubSpot "
                      f"owner(s) have no SF_Username: {names}{more}.",
                      "Fill SF_Username on the _Users tab for every active owner.")

    def run(self, users):
        self.check_holds()
        self.check_object_consistency()
        self.check_api_names()
        self.check_types()
        self.check_picklists()
        self.check_relationships()
        self.check_validation_rules()
        self.check_required_fields()
        self.check_owner_mapping(users)
        return self


def read_users() -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(WB, data_only=True)
    try:
        if "_Users" not in wb.sheetnames:
            return []
        ws = wb["_Users"]
        header = [ws.cell(row=5, column=c).value for c in range(2, 9)]
        users = []
        for r in range(6, ws.max_row + 1):
            values = [ws.cell(row=r, column=c).value for c in range(2, 9)]
            if not any(values):
                continue
            users.append({header[i]: (values[i] or "") for i in range(len(header)) if header[i]})
        return users
    finally:
        wb.close()


def summarise_deployables(tabs):
    print("\nWHAT WOULD DEPLOY")
    print("-" * 70)
    totals = {"Y": 0, "N": 0}
    for tab_name, tab in sorted(tabs.items()):
        parts = []
        for section_name in wbmod.SECTIONS:
            section = tab.sections.get(section_name)
            if not section or not section.rows:
                continue
            yes = sum(1 for r in section.rows
                      if str(r.get("Deploy", "")).strip().upper() == "Y")
            no = sum(1 for r in section.rows
                     if str(r.get("Deploy", "")).strip().upper() == "N")
            totals["Y"] += yes
            totals["N"] += no
            if yes or no:
                parts.append(f"{section_name.lower()} {yes}Y/{no}N")
        if parts:
            print(f"  {tab_name}: " + ", ".join(parts))
    print(f"\n  TOTAL: {totals['Y']} rows to deploy, {totals['N']} skipped by design.")


def approve(tabs, by: str):
    rows = list(wbmod.all_rows(tabs))
    digest = hashing.content_hash(rows)
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    wbmod.write_approval(WB, status="APPROVED", approved_by=by,
                         approved_at=now, content_hash=digest)
    print("\n" + "=" * 70)
    print("APPROVED")
    print("=" * 70)
    print(f"  by:   {by}")
    print(f"  at:   {now}")
    print(f"  hash: {digest}")
    print("\nAny edit to Final_*, Deploy or Notes from now on changes this hash,")
    print("and guard_deploy.py will block the deploy until you re-approve.")


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate the reviewed workbook")
    parser.add_argument("--approve", action="store_true",
                        help="on a clean pass, set _Approval to APPROVED and record the hash")
    parser.add_argument("--by", default="", help="name recorded as the approver")
    args = parser.parse_args()

    if not WB.exists():
        print(f"ERROR: {WB.relative_to(ROOT)} not found. Run 02_build_workbook.py first.")
        return 2

    try:
        tabs = wbmod.read_workbook(WB)
    except wbmod.WorkbookFormatError as exc:
        print(f"ERROR: workbook structure is damaged.\n  {exc}")
        print("  A marker row in hidden column A was probably deleted. Regenerate with 02.")
        return 2

    config = mapping.load_config()
    users = read_users()
    validator = Validator(tabs, config).run(users)

    print("=" * 70)
    print("WORKBOOK VALIDATION")
    print("=" * 70)

    if validator.failures:
        print(f"\n{len(validator.failures)} FAILURE(S) — nothing can deploy:\n")
        by_tab: dict[str, list[Issue]] = {}
        for issue in validator.failures:
            by_tab.setdefault(issue.tab, []).append(issue)
        for tab in sorted(by_tab):
            print(f"  [{tab}]")
            for issue in by_tab[tab][:20]:
                print(f"    row {issue.row}: {issue.problem}")
                print(f"      FIX: {issue.fix}")
            if len(by_tab[tab]) > 20:
                print(f"    … +{len(by_tab[tab]) - 20} more on this tab")
            print()

    if validator.warnings:
        print(f"{len(validator.warnings)} WARNING(S) — will not block a deploy:\n")
        for issue in validator.warnings[:15]:
            print(f"  {issue.tab}!row {issue.row}: {issue.problem}")
            print(f"    SUGGEST: {issue.fix}")
        if len(validator.warnings) > 15:
            print(f"  … +{len(validator.warnings) - 15} more")
        print()

    if validator.failures:
        print("=" * 70)
        print("RESULT: FAILED. Fix the above, then re-run.")
        print("=" * 70)
        return 1

    print("RESULT: PASSED" + (f" with {len(validator.warnings)} warning(s)."
                              if validator.warnings else "."))
    summarise_deployables(tabs)

    current = hashing.content_hash(list(wbmod.all_rows(tabs)))
    approval = wbmod.read_approval(WB)
    print(f"\n  Approval status: {approval['Status']}")
    print(f"  Current content hash: {hashing.short_hash(current)}…")
    if approval["Content_Hash"]:
        match = hashing.hashes_match(approval["Content_Hash"], current)
        print(f"  Recorded hash:        {hashing.short_hash(approval['Content_Hash'])}… "
              f"{'MATCHES' if match else 'STALE — the sheet changed after approval'}")

    if args.approve:
        if not args.by:
            print("\nERROR: --approve requires --by \"Your Name\".")
            return 1
        approve(tabs, args.by)
    else:
        print("\nSTOPPING HERE. This script never deploys.")
        print("To approve: python scripts/03_validate_sheet.py --approve --by \"Your Name\"")

    return 0


if __name__ == "__main__":
    sys.exit(main())
