#!/usr/bin/env python3
"""02 — Build the reviewable migration workbook.

    python scripts/02_build_workbook.py
    python scripts/02_build_workbook.py --preserve-decisions

Reads raw/ + config/, emits mapping/workbook.xlsx. One tab per object, sections
stacked inside, hidden column A carrying the markers that make the file
parseable back regardless of row positions.

Every Deploy cell defaults to HOLD, by design. 03 refuses to run while any HOLD
remains, so nothing can deploy that nobody has looked at.

CHECKPOINT: open the xlsx. Sections collapse. Dropdowns work. You cannot type
into a Proposed_* cell. Every Deploy cell says HOLD.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from lib import mapping, workbook as wbmod
from lib.mapping import Proposal

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "raw"
MAPPING = ROOT / "mapping"
OUT = MAPPING / "workbook.xlsx"

# -- styling -----------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
SECTION_FONT = Font(bold=True, size=11, color="1F3864")
PROPOSED_FILL = PatternFill("solid", fgColor="F2F2F2")   # protected
EDITABLE_FILL = PatternFill("solid", fgColor="FFF9E6")   # reviewer types here
FLAG_FONT = Font(color="C00000", size=9)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TAB_RED = "C00000"      # needs work
TAB_AMBER = "ED7D31"    # partially reviewed
TAB_GREEN = "70AD47"    # fully reviewed

MARKER_FONT = Font(color="FFFFFF", size=1)  # column A is hidden anyway


def log(msg=""):
    print(msg, flush=True)


def load_raw(name, default=None):
    path = RAW / name
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Sheet construction
# ---------------------------------------------------------------------------

class TabBuilder:
    """Writes one object tab, tracking row numbers for grouping and protection."""

    def __init__(self, ws, config):
        self.ws = ws
        self.config = config
        self.row = 1
        self.groups: list[tuple[int, int]] = []
        self.validations: dict[str, DataValidation] = {}

    def _marker(self, text):
        cell = self.ws.cell(row=self.row, column=1, value=text)
        cell.font = MARKER_FONT

    def section(self, name, title):
        self._marker(f"{wbmod.M_SECTION}{name}")
        cell = self.ws.cell(row=self.row, column=2, value=title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        self.ws.cell(row=self.row, column=2).alignment = Alignment(vertical="center")
        self.ws.row_dimensions[self.row].height = 20
        start = self.row
        self.row += 1
        return start

    def header(self, columns):
        self._marker(wbmod.M_HEADER)
        for i, name in enumerate(columns):
            cell = self.ws.cell(row=self.row, column=2 + i, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        self.ws.row_dimensions[self.row].height = 30
        self.row += 1

    def data_row(self, columns, values: dict):
        self._marker(wbmod.M_ROW)
        for i, name in enumerate(columns):
            value = values.get(name, "")
            cell = self.ws.cell(row=self.row, column=2 + i, value=value)
            cell.border = BORDER
            cell.font = Font(size=10)
            if wbmod.is_editable(name):
                cell.fill = EDITABLE_FILL
                cell.protection = Protection(locked=False)
            else:
                cell.fill = PROPOSED_FILL
                cell.protection = Protection(locked=True)
            if name == "Flags" and value:
                cell.font = FLAG_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if name in ("Notes", "Sample", "Proposed_Formula", "Final_Formula"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        self.row += 1

    def end(self, start_row):
        self._marker(wbmod.M_END)
        self.row += 1
        # Group the rows between the section header and #END so sections collapse.
        if self.row - 2 > start_row + 1:
            self.groups.append((start_row + 1, self.row - 2))
        self.row += 1  # blank spacer between sections

    def finish(self, columns_widths=None):
        for start, end in self.groups:
            self.ws.row_dimensions.group(start, end, outline_level=1, hidden=False)
        self.ws.column_dimensions["A"].hidden = True
        self.ws.column_dimensions["A"].width = 2
        for i, width in enumerate(columns_widths or []):
            self.ws.column_dimensions[get_column_letter(2 + i)].width = width
        self.ws.freeze_panes = "B2"
        self.ws.sheet_view.showGridLines = False
        # Protect the sheet, but leave editable cells unlocked. No password:
        # this guards against accidents, not against a determined reviewer.
        self.ws.protection.sheet = True
        self.ws.protection.enable()
        self.ws.protection.formatCells = False
        self.ws.protection.insertRows = True
        self.ws.protection.sort = True
        self.ws.protection.autoFilter = True


def add_validations(ws, config, ranges):
    """Dropdowns on Deploy and Final_Type."""
    deploy_dv = DataValidation(
        type="list", formula1='"Y,N,HOLD"', allow_blank=False,
        showErrorMessage=True, errorTitle="Invalid Deploy value",
        error="Must be Y, N, or HOLD.",
        promptTitle="Deploy?", prompt="Y = deploy this run. N = skip, by design. HOLD = undecided.",
    )
    ws.add_data_validation(deploy_dv)
    for ref in ranges.get("Deploy", []):
        deploy_dv.add(ref)

    types = ",".join(config["types"]["sf_field_types"])
    type_dv = DataValidation(
        type="list", formula1=f'"{types}"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid Salesforce field type",
        error="Pick a supported Salesforce field type.",
    )
    ws.add_data_validation(type_dv)
    for ref in ranges.get("Final_Type", []):
        type_dv.add(ref)

    # HOLD stands out in red until someone decides.
    for ref in ranges.get("Deploy", []):
        ws.conditional_formatting.add(ref, CellIsRule(
            operator="equal", formula=['"HOLD"'],
            fill=PatternFill("solid", bgColor="FFC7CE"),
            font=Font(color="9C0006", bold=True)))
        ws.conditional_formatting.add(ref, CellIsRule(
            operator="equal", formula=['"Y"'],
            fill=PatternFill("solid", bgColor="C6EFCE"),
            font=Font(color="006100")))


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def build(preserve: bool) -> int:
    if not (RAW / "schemas.json").exists():
        log("ERROR: raw/schemas.json not found. Run 01_fetch_hubspot.py first.")
        return 1

    _warn_if_stale()
    config = mapping.load_config()

    prior = {}
    if preserve and OUT.exists():
        prior = _read_prior_decisions()
        log(f"Preserving {len(prior)} prior decisions from the existing workbook.\n")

    schemas = load_raw("schemas.json", [])
    pipelines = load_raw("pipelines.json", {})
    owners = load_raw("owners.json", [])
    associations = load_raw("associations.json", {})
    samples = load_raw("samples.json", {})

    wb = Workbook()
    wb.remove(wb.active)

    index_rows = []
    stats = {"objects": 0, "custom_fields": 0, "standard_maps": 0,
             "flagged": 0, "holds": 0, "picklists": 0, "record_types": 0,
             "relationships": 0}
    all_flags: list[tuple[str, str, str]] = []

    obj_config = config["objects"]

    for schema in schemas:
        hs_object = schema.get("name")
        if not hs_object:
            continue
        if hs_object in {e.get("object") for e in (obj_config.get("excluded") or [])}:
            continue

        props = load_raw(f"properties_{hs_object}.json", [])
        if not props:
            continue

        entry = (obj_config.get("objects") or {}).get(hs_object)
        sf_object, is_custom = _resolve_sf_object(hs_object, schema, entry, config)

        tab_name = _tab_name(sf_object)
        ws = wb.create_sheet(tab_name)
        builder = TabBuilder(ws, config)
        ranges = {"Deploy": [], "Final_Type": []}

        _title_block(ws, builder, hs_object, sf_object, is_custom, schema)

        proposals = _build_proposals(hs_object, props, config, samples.get(hs_object, []))
        mapping.detect_collisions(proposals)

        counts = _write_fields(builder, proposals, ranges, prior, tab_name)
        stats["custom_fields"] += counts["custom"]
        stats["standard_maps"] += counts["standard"]
        stats["flagged"] += counts["flagged"]

        for p in proposals:
            for flag in p.flags:
                all_flags.append((tab_name, p.hs_name, flag))

        stats["picklists"] += _write_picklists(builder, proposals, ranges, prior, tab_name)
        stats["relationships"] += _write_relationships(
            builder, hs_object, associations, config, ranges, prior, tab_name)

        if entry is None or entry.get("record_types") in (True, "auto"):
            stats["record_types"] += _write_record_types(
                builder, hs_object, sf_object, pipelines, config, ranges, prior, tab_name)

        _write_validation_rules(builder, ranges, prior, tab_name)

        builder.finish(_widths())
        add_validations(ws, config, ranges)

        holds = _count_holds(ws)
        stats["holds"] += holds
        ws.sheet_properties.tabColor = (
            TAB_RED if holds else TAB_GREEN)

        index_rows.append({
            "HubSpot Object": hs_object,
            "Salesforce Object": sf_object,
            "Type": "Custom" if is_custom else "Standard",
            "Properties": len(props),
            "New Fields": counts["custom"],
            "Standard Maps": counts["standard"],
            "Flagged": counts["flagged"],
            "HOLD": holds,
            "Tab": tab_name,
        })
        stats["objects"] += 1
        log(f"  {tab_name}: {counts['custom']} new fields, "
            f"{counts['standard']} standard maps, {counts['flagged']} flagged, "
            f"{holds} HOLD")

    _build_index(wb, index_rows, stats)
    _build_users(wb, owners, prior)
    _build_approval(wb)
    _build_runlog(wb)

    wb.move_sheet("_Index", offset=-len(wb.sheetnames))
    MAPPING.mkdir(exist_ok=True)
    wb.save(OUT)

    _report(stats, all_flags)
    return 0


def _resolve_sf_object(hs_object, schema, entry, config):
    if entry:
        return entry["sf_object"], entry.get("sf_type") == "custom"
    # Unknown object → the defaults path. Expected for every custom object,
    # since the portal contents were unknown at build time.
    naming = config["naming"]
    pascal, _ = mapping.to_api_name(hs_object, naming, suffix=False)
    return f"{pascal}__c", True


def _tab_name(sf_object: str) -> str:
    """Excel tab names cap at 31 chars and reject : \\ / ? * [ ]"""
    name = re.sub(r"[:\\/?*\[\]]", "_", sf_object)
    return name[:31]


def _title_block(ws, builder, hs_object, sf_object, is_custom, schema):
    ws.cell(row=builder.row, column=2, value=f"{hs_object}  →  {sf_object}").font = Font(
        bold=True, size=14, color="1F3864")
    builder.row += 1
    kind = "custom object (will be created)" if is_custom else "standard object (exists in org)"
    label = schema.get("labels", {}).get("singular", "")
    ws.cell(row=builder.row, column=2,
            value=f"{kind} · HubSpot label: {label}").font = Font(size=9, italic=True,
                                                                  color="595959")
    builder.row += 1
    ws.cell(row=builder.row, column=2,
            value="Edit only the yellow columns: Final_*, Deploy, Notes. "
                  "Set every Deploy cell to Y or N — leave nothing at HOLD."
            ).font = Font(size=9, italic=True, color="C00000")
    builder.row += 2


def _build_proposals(hs_object, props, config, records) -> list[Proposal]:
    names = [p.get("name", "") for p in props]
    fills = mapping.compute_fill_rates(records, names)
    max_lens = mapping.observed_max_lengths(records)

    proposals = []
    for prop in props:
        name = prop.get("name", "")
        proposals.append(mapping.map_property(
            prop, hs_object, config,
            fill_percent=fills.get(name),
            observed_max_len=max_lens.get(name, 0),
            sample=mapping.first_sample(records, name),
        ))
    return proposals


def _decision_key(tab, section, key):
    return f"{tab}|{section}|{key}"


def _apply_prior(values, prior, tab, section, key):
    """Carry a reviewer's earlier decision forward across a regeneration."""
    saved = prior.get(_decision_key(tab, section, key))
    if saved:
        for column, value in saved.items():
            if value not in (None, ""):
                values[column] = value


def _write_fields(builder, proposals, ranges, prior, tab_name) -> dict:
    columns = wbmod.FIELD_COLUMNS
    start = builder.section("FIELDS", "FIELDS")
    builder.header(columns)
    first_data_row = builder.row
    counts = {"custom": 0, "standard": 0, "flagged": 0}

    for p in proposals:
        values = {
            "HS Property": p.hs_name,
            "HS Type": p.hs_type,
            "HS FieldType": p.hs_field_type,
            "Fill %": f"{p.fill_percent:.1f}%" if p.fill_percent is not None else "",
            "Sample": p.sample,
            "Proposed_API": p.api_name,
            "Proposed_Type": p.sf_type,
            "Proposed_Len": p.length or p.precision or "",
            "Final_API": p.api_name,
            "Final_Type": p.sf_type,
            "Final_Len": p.length or p.precision or "",
            "Deploy": p.deploy_default,
            "Notes": "",
            "Flags": " | ".join(p.flags),
        }
        _apply_prior(values, prior, tab_name, "FIELDS", p.hs_name)
        builder.data_row(columns, values)

        if p.is_standard:
            counts["standard"] += 1
        else:
            counts["custom"] += 1
        if p.flags:
            counts["flagged"] += 1

    _record_ranges(ranges, columns, first_data_row, builder.row - 1)
    builder.end(start)
    return counts


def _write_picklists(builder, proposals, ranges, prior, tab_name) -> int:
    picklists = [p for p in proposals if p.picklist_values]
    if not picklists:
        return 0

    columns = wbmod.PICKLIST_COLUMNS
    start = builder.section("PICKLISTS", "PICKLIST VALUES")
    builder.header(columns)
    first = builder.row
    count = 0

    for p in picklists:
        for v in p.picklist_values:
            values = {
                "HS Property": p.hs_name,
                "HS Value": v["fullName"],
                "HS Label": v["label"],
                "Sort": v.get("sort", 0),
                "Proposed_Value": v["fullName"],
                "Final_Value": v["fullName"],
                "Default": "",
                "Deploy": "HOLD",
                "Notes": "",
                "Flags": "",
            }
            _apply_prior(values, prior, tab_name, "PICKLISTS",
                         f"{p.hs_name}:{v['fullName']}")
            builder.data_row(columns, values)
            count += 1

    _record_ranges(ranges, columns, first, builder.row - 1)
    builder.end(start)
    return count


def _write_relationships(builder, hs_object, associations, config, ranges,
                         prior, tab_name) -> int:
    pairs = {k: v for k, v in (associations or {}).items()
             if k.startswith(f"{hs_object}__to__")}
    if not pairs:
        return 0

    columns = wbmod.RELATIONSHIP_COLUMNS
    start = builder.section("RELATIONSHIPS", "RELATIONSHIPS")
    builder.header(columns)
    first = builder.row
    naming = config["naming"]
    obj_config = config["objects"]
    count = 0

    for key, labels in pairs.items():
        to_object = key.split("__to__")[1]
        entry = (obj_config.get("objects") or {}).get(to_object)
        target = entry["sf_object"] if entry else _resolve_sf_object(
            to_object, {}, None, config)[0]

        user_defined = [l for l in labels if l.get("category") == "USER_DEFINED"]
        cardinality = "many-to-many" if len(user_defined) > 1 else "many-to-one"

        api_name, flags = mapping.to_api_name(f"{to_object}", naming)
        if cardinality == "many-to-many":
            flags.append(
                f"{len(user_defined)} labelled association types — HubSpot "
                "many-to-many has no direct lookup equivalent. Needs a junction "
                f"object ({hs_object}_{to_object}_Assoc__c) or a chosen primary."
            )

        values = {
            "HS Association": key,
            "From Object": hs_object,
            "To Object": to_object,
            "Cardinality": cardinality,
            "Proposed_API": api_name,
            "Proposed_Type": "Lookup",  # never propose MasterDetail automatically
            "Proposed_Target": target,
            "Child_Relationship": api_name.replace("__c", ""),
            "Final_API": api_name,
            "Final_Type": "Lookup",
            "Final_Target": target,
            "Deploy": "HOLD",
            "Notes": "",
            "Flags": " | ".join(flags),
        }
        _apply_prior(values, prior, tab_name, "RELATIONSHIPS", key)
        builder.data_row(columns, values)
        count += 1

    _record_ranges(ranges, columns, first, builder.row - 1)
    builder.end(start)
    return count


def _write_record_types(builder, hs_object, sf_object, pipelines, config,
                        ranges, prior, tab_name) -> int:
    found = (pipelines or {}).get(hs_object) or []
    if not found:
        return 0

    columns = wbmod.RECORD_TYPE_COLUMNS
    start = builder.section("RECORD_TYPES", "RECORD TYPES / PIPELINES")
    builder.header(columns)
    first = builder.row
    naming = config["naming"]
    process_suffix = naming["components"]["record_type"]["sales_process_template"]
    count = 0

    for pipeline in found:
        rt_api, rt_flags = mapping.to_api_name(pipeline.get("label", ""), naming, suffix=False)
        process = process_suffix.replace("{RecordType}", rt_api)

        for stage in pipeline.get("stages", []):
            stage_value, stage_flags = mapping.to_api_name(
                stage.get("label", ""), naming, suffix=False)
            probability = (stage.get("metadata") or {}).get("probability", "")

            values = {
                "HS Pipeline": pipeline.get("label", ""),
                "HS Stage": stage.get("label", ""),
                "Stage Order": stage.get("displayOrder", 0),
                "Probability": probability,
                "Proposed_RT_API": rt_api,
                "Proposed_RT_Label": pipeline.get("label", ""),
                "Proposed_Stage_Value": stage.get("label", ""),
                "Process_Name": process,
                "Final_RT_API": rt_api,
                "Final_RT_Label": pipeline.get("label", ""),
                "Final_Stage_Value": stage.get("label", ""),
                "Deploy": "HOLD",
                "Notes": "",
                "Flags": " | ".join(rt_flags + stage_flags),
            }
            _apply_prior(values, prior, tab_name, "RECORD_TYPES",
                         f"{pipeline.get('label')}:{stage.get('label')}")
            builder.data_row(columns, values)
            count += 1

    _record_ranges(ranges, columns, first, builder.row - 1)
    builder.end(start)
    return count


def _write_validation_rules(builder, ranges, prior, tab_name):
    """Empty by design — rules are authored by the consultant during review,
    not derived from HubSpot. The section exists so there is somewhere to put
    them, and so 03/04 have a consistent shape to read."""
    columns = wbmod.VALIDATION_RULE_COLUMNS
    start = builder.section("VALIDATION_RULES", "VALIDATION RULES  (add rows below as needed)")
    builder.header(columns)
    first = builder.row
    builder.data_row(columns, {
        "Rule_Name": "", "Source": "manual", "Description": "",
        "Proposed_Formula": "", "Final_Formula": "",
        "Error_Message_Text": "", "Error_Location": "",
        "Active": "FALSE", "Deploy": "N", "Notes": "",
        "Flags": "Rules deploy inactive or after the data load — active rules "
                 "reject legitimate historical records.",
    })
    _record_ranges(ranges, columns, first, builder.row - 1)
    builder.end(start)


def _record_ranges(ranges, columns, first_row, last_row):
    if last_row < first_row:
        return
    for name in ("Deploy", "Final_Type"):
        if name in columns:
            letter = get_column_letter(2 + columns.index(name))
            ranges.setdefault(name, []).append(f"{letter}{first_row}:{letter}{last_row}")


def _widths():
    return [22, 12, 14, 9, 26, 26, 16, 11, 26, 16, 11, 9, 30, 46,
            18, 20, 34, 10, 40]


def _count_holds(ws) -> int:
    count = 0
    header = None
    for r in range(1, ws.max_row + 1):
        marker = ws.cell(row=r, column=1).value
        if marker == wbmod.M_HEADER:
            header = [ws.cell(row=r, column=c).value for c in range(2, ws.max_column + 1)]
        elif marker == wbmod.M_ROW and header and "Deploy" in header:
            col = 2 + header.index("Deploy")
            if str(ws.cell(row=r, column=col).value or "").strip().upper() == "HOLD":
                count += 1
    return count


# ---------------------------------------------------------------------------
# Control tabs
# ---------------------------------------------------------------------------

def _build_index(wb, rows, stats):
    ws = wb.create_sheet("_Index")
    ws.sheet_properties.tabColor = "1F3864"

    ws["B2"] = "HubSpot → Salesforce Migration Workbook"
    ws["B2"].font = Font(bold=True, size=16, color="1F3864")
    ws["B3"] = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["B3"].font = Font(size=9, italic=True, color="595959")
    ws["B4"] = ("Review instructions: edit only Final_*, Deploy and Notes. "
                "Set every Deploy cell to Y or N. Leave nothing at HOLD.")
    ws["B4"].font = Font(size=10, bold=True, color="C00000")

    headers = ["HubSpot Object", "Salesforce Object", "Type", "Properties",
               "New Fields", "Standard Maps", "Flagged", "HOLD", "Go to"]
    for i, name in enumerate(headers):
        cell = ws.cell(row=6, column=2 + i, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r, row in enumerate(rows, start=7):
        for i, key in enumerate(headers[:-1]):
            cell = ws.cell(row=r, column=2 + i, value=row.get(key, ""))
            cell.border = BORDER
            if key == "HOLD" and row.get("HOLD"):
                cell.font = Font(color="C00000", bold=True)
            elif key == "Flagged" and row.get("Flagged"):
                cell.font = Font(color="ED7D31", bold=True)
        link = ws.cell(row=r, column=2 + len(headers) - 1)
        link.value = f'=HYPERLINK("#\'{row["Tab"]}\'!B1","open {row["Tab"]} →")'
        link.font = Font(color="0563C1", underline="single")
        link.border = BORDER

    summary_row = 7 + len(rows) + 2
    ws.cell(row=summary_row, column=2, value="TOTALS").font = Font(bold=True, size=12)
    for i, (label, value) in enumerate([
        ("Objects", stats["objects"]),
        ("New custom fields", stats["custom_fields"]),
        ("Standard field maps", stats["standard_maps"]),
        ("Picklist values", stats["picklists"]),
        ("Record type stages", stats["record_types"]),
        ("Relationships", stats["relationships"]),
        ("Flagged for review", stats["flagged"]),
        ("Still at HOLD", stats["holds"]),
    ], start=1):
        ws.cell(row=summary_row + i, column=2, value=label).font = Font(size=10)
        cell = ws.cell(row=summary_row + i, column=3, value=value)
        cell.font = Font(bold=True, size=10,
                         color="C00000" if label == "Still at HOLD" and value else "000000")

    for col, width in zip("BCDEFGHIJ", [24, 26, 10, 12, 12, 15, 10, 10, 22]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


def _build_users(wb, owners, prior):
    """HubSpot owner → Salesforce user. Unmapped owners silently default
    ownership to the running user, so 03 hard-fails on any blank."""
    ws = wb.create_sheet("_Users")
    ws.sheet_properties.tabColor = "7030A0"

    ws["B2"] = "OWNER MAPPING"
    ws["B2"].font = Font(bold=True, size=14, color="1F3864")
    ws["B3"] = ("Every active owner needs a Salesforce username. Unmapped owners "
                "silently default record ownership to whoever runs the load.")
    ws["B3"].font = Font(size=9, italic=True, color="C00000")

    headers = ["HS Owner Id", "HS Email", "HS Name", "Archived",
               "SF_Username", "SF_User_Id", "Notes"]
    for i, name in enumerate(headers):
        cell = ws.cell(row=5, column=2 + i, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for r, owner in enumerate(owners or [], start=6):
        full_name = f"{owner.get('firstName', '')} {owner.get('lastName', '')}".strip()
        flags = []
        if owner.get("archived"):
            flags.append("ARCHIVED in HubSpot — map to a fallback user or leave blank and set N.")
        if not owner.get("email"):
            flags.append("No email — cannot auto-match to a Salesforce user.")

        values = [owner.get("id", ""), owner.get("email", ""), full_name,
                  "Y" if owner.get("archived") else "N", "", "", " | ".join(flags)]
        for i, value in enumerate(values):
            cell = ws.cell(row=r, column=2 + i, value=value)
            cell.border = BORDER
            if headers[i].startswith("SF_"):
                cell.fill = EDITABLE_FILL
                cell.protection = Protection(locked=False)
            if headers[i] == "Notes" and value:
                cell.font = FLAG_FONT

    for col, width in zip("BCDEFGH", [14, 32, 24, 10, 32, 20, 44]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


def _build_approval(wb):
    """Exactly four cells. The hash is what guard_deploy.py checks."""
    ws = wb.create_sheet("_Approval")
    ws.sheet_properties.tabColor = "C00000"

    ws["A1"] = "Status"
    ws["B1"] = "DRAFT"
    ws["A2"] = "Approved_By"
    ws["B2"] = ""
    ws["A3"] = "Approved_At"
    ws["B3"] = ""
    ws["A4"] = "Content_Hash"
    ws["B4"] = ""

    for r in range(1, 5):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).fill = EDITABLE_FILL

    ws["A6"] = "Set Status to APPROVED only via: python scripts/03_validate_sheet.py --approve"
    ws["A6"].font = Font(size=9, italic=True, color="595959")
    ws["A7"] = ("The hash covers Final_*, Deploy and Notes. If the sheet changes after "
                "approval the hash stops matching and guard_deploy.py blocks the deploy.")
    ws["A7"].font = Font(size=9, italic=True, color="595959")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 72
    ws.sheet_view.showGridLines = False


def _build_runlog(wb):
    ws = wb.create_sheet("_RunLog")
    ws.sheet_properties.tabColor = "808080"
    headers = ["Timestamp", "Script", "Org", "Action", "Result", "Detail"]
    for i, name in enumerate(headers):
        cell = ws.cell(row=1, column=1 + i, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for col, width in zip("ABCDEF", [22, 22, 16, 20, 12, 70]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _warn_if_stale():
    schemas = RAW / "schemas.json"
    if not schemas.exists():
        return
    age_days = (datetime.now().timestamp() - schemas.stat().st_mtime) / 86400
    if age_days > 7:
        log(f"WARNING: raw/ is {age_days:.0f} days old. A stale schema produces a "
            "workbook that reviews cleanly and deploys wrong. Re-run 01.\n")


def _read_prior_decisions() -> dict:
    try:
        tabs = wbmod.read_workbook(OUT)
    except Exception as exc:
        log(f"WARNING: could not read prior decisions ({exc}). Starting fresh.")
        return {}

    prior = {}
    for row in wbmod.all_rows(tabs):
        key = (row.get("HS Property") or row.get("HS Association")
               or row.get("Rule_Name") or "")
        if row.get("_section") == "PICKLISTS":
            key = f"{row.get('HS Property')}:{row.get('HS Value')}"
        elif row.get("_section") == "RECORD_TYPES":
            key = f"{row.get('HS Pipeline')}:{row.get('HS Stage')}"
        if not key:
            continue
        saved = {k: v for k, v in row.items()
                 if (k.startswith("Final_") or k in ("Deploy", "Notes", "Active", "Default"))
                 and v not in (None, "")}
        if saved:
            prior[_decision_key(row["_tab"], row["_section"], key)] = saved
    return prior


def _report(stats, all_flags):
    log("\n" + "=" * 70)
    log(f"WORKBOOK: {OUT.relative_to(ROOT)}")
    log("=" * 70)
    log(f"  {stats['objects']} objects")
    log(f"  {stats['custom_fields']} new custom fields proposed")
    log(f"  {stats['standard_maps']} properties mapped to existing standard fields")
    log(f"  {stats['picklists']} picklist values")
    log(f"  {stats['record_types']} record type / stage rows")
    log(f"  {stats['relationships']} relationships")
    log(f"  {stats['flagged']} rows flagged for review")
    log(f"  {stats['holds']} rows still at HOLD  ← all of them, by design")

    if all_flags:
        log(f"\nFLAGGED ITEMS — look at these first ({len(all_flags)}):")
        by_kind: dict[str, list[str]] = {}
        for tab, prop, flag in all_flags:
            kind = flag.split(":")[0].split("—")[0].strip()[:60]
            by_kind.setdefault(kind, []).append(f"{tab}.{prop}")
        for kind, items in sorted(by_kind.items(), key=lambda kv: -len(kv[1])):
            log(f"\n  [{len(items)}] {kind}")
            for item in items[:6]:
                log(f"      - {item}")
            if len(items) > 6:
                log(f"      … +{len(items) - 6} more")

    log("\nCHECKPOINT — open the workbook and confirm:")
    log("  1. Sections collapse (the +/- outline in the left margin)")
    log("  2. The Deploy dropdown offers Y / N / HOLD")
    log("  3. Typing into a Proposed_* cell is refused")
    log("  4. Every Deploy cell reads HOLD")
    log("\nNEXT: review, then python scripts/03_validate_sheet.py")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the migration workbook")
    parser.add_argument("--preserve-decisions", action="store_true",
                        help="carry existing Final_*/Deploy/Notes values forward")
    args = parser.parse_args()
    return build(preserve=args.preserve_decisions)


if __name__ == "__main__":
    sys.exit(main())
