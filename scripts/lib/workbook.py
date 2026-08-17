"""Section-marker read/write for the migration workbook.

The workbook must survive human editing: inserted rows, deleted rows, sorted
ranges, added comments. So nothing is addressed by row number. Column A holds a
hidden marker that names what each row is, and the file is parsed back by
scanning those markers.

    A                    B ...
    #SECTION:FIELDS
    #HEADER              HS Property | HS Type | ...
    #ROW                 firstname   | string  | ...
    #END

Column A is hidden in the UI, and protected, so reviewers neither see nor break
it. If a reviewer deletes a marker row anyway, the parser fails loudly on that
tab rather than silently mis-reading the rest.
"""

from __future__ import annotations

from dataclasses import dataclass, field

MARKER_COL = 1  # column A

M_SECTION = "#SECTION:"
M_HEADER = "#HEADER"
M_ROW = "#ROW"
M_END = "#END"

CONTROL_TABS = ("_Index", "_Users", "_Approval", "_RunLog")

SECTIONS = (
    "FIELDS",
    "PICKLISTS",
    "RELATIONSHIPS",
    "RECORD_TYPES",
    "VALIDATION_RULES",
)

# Column layouts per section. Order is authoritative — it defines the sheet.
FIELD_COLUMNS = [
    "HS Property", "HS Type", "HS FieldType", "Fill %", "Sample",
    "Proposed_API", "Proposed_Type", "Proposed_Len",
    "Final_API", "Final_Type", "Final_Len",
    "Deploy", "Notes", "Flags",
    "Deployed_At", "Component_Id", "Setup_Link", "Result", "Error",
]

PICKLIST_COLUMNS = [
    "HS Property", "HS Value", "HS Label", "Sort",
    "Proposed_Value", "Final_Value", "Default",
    "Deploy", "Notes", "Flags",
    "Deployed_At", "Result", "Error",
]

RELATIONSHIP_COLUMNS = [
    "HS Association", "From Object", "To Object", "Cardinality",
    "Proposed_API", "Proposed_Type", "Proposed_Target", "Child_Relationship",
    "Final_API", "Final_Type", "Final_Target",
    "Deploy", "Notes", "Flags",
    "Deployed_At", "Component_Id", "Setup_Link", "Result", "Error",
]

RECORD_TYPE_COLUMNS = [
    "HS Pipeline", "HS Stage", "Stage Order", "Probability",
    "Proposed_RT_API", "Proposed_RT_Label", "Proposed_Stage_Value", "Process_Name",
    "Final_RT_API", "Final_RT_Label", "Final_Stage_Value",
    "Deploy", "Notes", "Flags",
    "Deployed_At", "Component_Id", "Setup_Link", "Result", "Error",
]

VALIDATION_RULE_COLUMNS = [
    "Rule_Name", "Source", "Description",
    "Proposed_Formula", "Final_Formula", "Error_Message_Text", "Error_Location",
    "Active", "Deploy", "Notes", "Flags",
    "Deployed_At", "Component_Id", "Setup_Link", "Result", "Error",
]

SECTION_COLUMNS = {
    "FIELDS": FIELD_COLUMNS,
    "PICKLISTS": PICKLIST_COLUMNS,
    "RELATIONSHIPS": RELATIONSHIP_COLUMNS,
    "RECORD_TYPES": RECORD_TYPE_COLUMNS,
    "VALIDATION_RULES": VALIDATION_RULE_COLUMNS,
}

# Reviewers may edit these. Everything else is protected.
EDITABLE_PREFIXES = ("Final_",)
EDITABLE_EXACT = ("Deploy", "Notes", "Active", "Default")

DEPLOY_VALUES = ("Y", "N", "HOLD")


class WorkbookFormatError(Exception):
    """Raised when markers are missing or malformed — never guessed around."""


@dataclass
class Section:
    name: str
    header: list[str] = field(default_factory=list)
    rows: list[dict] = field(default_factory=list)
    header_row: int = 0
    row_numbers: list[int] = field(default_factory=list)


@dataclass
class Tab:
    name: str
    sections: dict[str, Section] = field(default_factory=dict)


def is_editable(column: str) -> bool:
    return column.startswith(EDITABLE_PREFIXES) or column in EDITABLE_EXACT


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def read_workbook(path):
    """Parse a workbook into {tab_name: Tab}. Control tabs are skipped."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    tabs = {}
    for ws in wb.worksheets:
        if ws.title in CONTROL_TABS:
            continue
        tabs[ws.title] = _read_tab(ws)
    wb.close()
    return tabs


def _read_tab(ws) -> Tab:
    tab = Tab(name=ws.title)
    current = None
    header = None

    for r in range(1, ws.max_row + 1):
        marker = ws.cell(row=r, column=MARKER_COL).value
        if marker is None:
            continue
        marker = str(marker).strip()

        if marker.startswith(M_SECTION):
            name = marker[len(M_SECTION):].strip()
            current = Section(name=name)
            tab.sections[name] = current
            header = None

        elif marker == M_HEADER:
            if current is None:
                raise WorkbookFormatError(
                    f"{ws.title} row {r}: #HEADER outside any #SECTION. "
                    "A marker row in column A was probably deleted."
                )
            header = _row_values(ws, r)
            current.header = header
            current.header_row = r

        elif marker == M_ROW:
            if current is None or header is None:
                raise WorkbookFormatError(
                    f"{ws.title} row {r}: #ROW with no preceding #HEADER."
                )
            values = _row_values(ws, r)
            row = {header[i]: (values[i] if i < len(values) else None)
                   for i in range(len(header))}
            row["_tab"] = ws.title
            row["_section"] = current.name
            row["_row"] = r
            current.rows.append(row)
            current.row_numbers.append(r)

        elif marker == M_END:
            current = None
            header = None

    return tab


def _row_values(ws, r) -> list:
    """Everything to the right of the marker column, trailing blanks trimmed."""
    values = [ws.cell(row=r, column=c).value for c in range(MARKER_COL + 1, ws.max_column + 1)]
    while values and values[-1] is None:
        values.pop()
    return [v if v is not None else "" for v in values]


def all_rows(tabs):
    """Flatten every data row across every tab and section."""
    for tab in tabs.values():
        for section in tab.sections.values():
            yield from section.rows


def rows_where(tabs, section=None, deploy=None):
    for row in all_rows(tabs):
        if section and row.get("_section") != section:
            continue
        if deploy and str(row.get("Deploy", "")).strip().upper() != deploy.upper():
            continue
        yield row


def holds(tabs):
    """Every row still at HOLD. Non-empty means nothing may deploy."""
    return [r for r in all_rows(tabs)
            if str(r.get("Deploy", "")).strip().upper() == "HOLD"]


# ---------------------------------------------------------------------------
# Approval tab
# ---------------------------------------------------------------------------

def read_approval(path) -> dict:
    """Read the four-cell _Approval tab. Missing tab is an unapproved state."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    try:
        if "_Approval" not in wb.sheetnames:
            return {"Status": "MISSING", "Approved_By": "", "Approved_At": "",
                    "Content_Hash": ""}
        ws = wb["_Approval"]
        data = {}
        for r in range(1, ws.max_row + 1):
            key = ws.cell(row=r, column=1).value
            if key:
                data[str(key).strip()] = ws.cell(row=r, column=2).value or ""
        return {
            "Status": str(data.get("Status", "DRAFT")).strip().upper(),
            "Approved_By": data.get("Approved_By", ""),
            "Approved_At": data.get("Approved_At", ""),
            "Content_Hash": str(data.get("Content_Hash", "")).strip(),
        }
    finally:
        wb.close()


def write_approval(path, status=None, approved_by=None, approved_at=None,
                   content_hash=None):
    """Update _Approval in place, leaving unspecified fields untouched."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["_Approval"] if "_Approval" in wb.sheetnames else wb.create_sheet("_Approval")

    updates = {
        "Status": status,
        "Approved_By": approved_by,
        "Approved_At": approved_at,
        "Content_Hash": content_hash,
    }
    existing = {}
    for r in range(1, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        if key:
            existing[str(key).strip()] = r

    for key, value in updates.items():
        if value is None:
            continue
        row = existing.get(key)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Writeback
# ---------------------------------------------------------------------------

def write_cells(path, updates):
    """Apply {(tab, row_number, column_name): value} without disturbing layout.

    Used by 06 to write deploy results back. Column position is resolved from
    the section header, so it stays correct even if columns move.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_cache = {}

    for (tab_name, row_no, column), value in updates.items():
        ws = wb[tab_name]
        key = (tab_name, row_no)
        if key not in header_cache:
            header_cache[key] = _header_for_row(ws, row_no)
        header = header_cache[key]
        if column not in header:
            continue  # column absent from this section — nothing to write
        col_index = MARKER_COL + 1 + header.index(column)
        ws.cell(row=row_no, column=col_index, value=value)

    wb.save(path)
    wb.close()


def _header_for_row(ws, row_no) -> list[str]:
    """Walk upward to the nearest #HEADER above this row."""
    for r in range(row_no - 1, 0, -1):
        marker = ws.cell(row=r, column=MARKER_COL).value
        if marker and str(marker).strip() == M_HEADER:
            return _row_values(ws, r)
    raise WorkbookFormatError(f"{ws.title} row {row_no}: no #HEADER above it.")
