#!/usr/bin/env python3
"""DEV TOOL — simulate a human review of the workbook.

Not part of the delivery pipeline. It exists so scripts 03-06 can be tested
end to end before a real reviewer exists, and so the deliberate-contradiction
checkpoint can be reproduced on demand.

    python scripts/dev_simulate_review.py                # sensible Y/N everywhere
    python scripts/dev_simulate_review.py --break-object # field Y under object N
    python scripts/dev_simulate_review.py --break-name   # invalid API name
    python scripts/dev_simulate_review.py --leave-hold 3 # leave N rows at HOLD

Never run this against a workbook a client has actually reviewed — it
overwrites Deploy decisions.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from lib import workbook as wbmod

ROOT = Path(__file__).resolve().parent.parent
WB = ROOT / "mapping" / "workbook.xlsx"

# Properties a reviewer would realistically decline.
DECLINE = {
    "legacy_crm_identifier",   # 3% filled
    "engagement_score",        # calculated
    "hs_forecast_amount",      # calculated
}


def decide(row) -> str:
    section = row.get("_section")
    prop = str(row.get("HS Property", "")).lower()

    if section == "FIELDS":
        if prop in DECLINE:
            return "N"
        if prop.startswith("hs_") and prop != "hs_object_id":
            return "N"
        return "Y"
    if section == "VALIDATION_RULES":
        return "N"   # none authored yet
    return "Y"


def main() -> int:
    parser = argparse.ArgumentParser(description="DEV: simulate a workbook review")
    parser.add_argument("--break-object", action="store_true",
                        help="mark every field on one tab N but leave a relationship Y")
    parser.add_argument("--break-name", action="store_true",
                        help="write an invalid API name into a deploying field")
    parser.add_argument("--leave-hold", type=int, default=0,
                        help="leave N rows at HOLD")
    args = parser.parse_args()

    if not WB.exists():
        print("ERROR: no workbook. Run 02_build_workbook.py first.")
        return 1

    tabs = wbmod.read_workbook(WB)
    updates = {}
    held = 0
    rows = list(wbmod.all_rows(tabs))

    for row in rows:
        if held < args.leave_hold:
            held += 1
            continue  # leave it at HOLD
        updates[(row["_tab"], row["_row"], "Deploy")] = decide(row)

    # Map every active owner, so the owner check passes.
    users_updates = _map_owners()

    if args.break_object:
        target = "Case"
        for row in rows:
            if row["_tab"] == target and row["_section"] == "FIELDS":
                updates[(row["_tab"], row["_row"], "Deploy")] = "N"
            elif row["_tab"] == target and row["_section"] == "RELATIONSHIPS":
                updates[(row["_tab"], row["_row"], "Deploy")] = "Y"
        print(f"BROKEN DELIBERATELY: every {target} field set to N, "
              "relationship left at Y.")

    if args.break_name:
        for row in rows:
            if row["_section"] == "FIELDS" and row["_tab"] == "Contact":
                if str(row.get("Proposed_Type")) != "Standard":
                    updates[(row["_tab"], row["_row"], "Final_API")] = "Order__c"
                    print(f"BROKEN DELIBERATELY: {row['_tab']}!row {row['_row']} "
                          "Final_API set to the reserved word Order__c.")
                    break

    wbmod.write_cells(WB, updates)
    if users_updates:
        _write_users(users_updates)

    counts = {"Y": 0, "N": 0}
    for value in updates.values():
        if value in counts:
            counts[value] += 1
    print(f"Simulated review: {counts['Y']} Y, {counts['N']} N, "
          f"{args.leave_hold} left at HOLD.")
    print("NEXT: python scripts/03_validate_sheet.py")
    return 0


def _map_owners():
    from openpyxl import load_workbook

    wb = load_workbook(WB)
    ws = wb["_Users"]
    updates = []
    for r in range(6, ws.max_row + 1):
        owner_id = ws.cell(row=r, column=2).value
        email = ws.cell(row=r, column=3).value
        archived = ws.cell(row=r, column=5).value
        if not owner_id:
            continue
        if archived == "Y":
            continue  # archived owners are legitimately left unmapped
        updates.append((r, email or f"user{owner_id}@example.com.sandbox"))
    wb.close()
    return updates


def _write_users(updates):
    from openpyxl import load_workbook

    wb = load_workbook(WB)
    ws = wb["_Users"]
    for row, username in updates:
        ws.cell(row=row, column=6, value=username)   # SF_Username
    wb.save(WB)
    wb.close()
    print(f"Mapped {len(updates)} active owners on _Users.")


if __name__ == "__main__":
    sys.exit(main())
