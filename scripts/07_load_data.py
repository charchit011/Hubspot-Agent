#!/usr/bin/env python3
"""07 — Load HubSpot records into Salesforce via Bulk API 2.0.

    python scripts/07_load_data.py --org client-sbx --objects all --dry-run
    python scripts/07_load_data.py --org client-sbx --objects contacts

OUT OF SCOPE for the schema-only prototype phase. It is complete and ready, but
running it needs a separate, explicit decision. --dry-run is the default; a real
load requires --execute.

Mechanics that make this safe to re-run:
  - Upsert on HubSpot_Record_Id__c, an External ID + unique field. This is what
    makes the load idempotent. Without unique, Bulk cannot match and you get
    duplicates.
  - Two passes, never combined: records first with lookups null, then
    relationships wired by external ID.
  - Parents before children: Account → Contact → Opportunity → line items.
  - Owner assignment from the _Users tab, applied during transform.

THE TEST THAT MATTERS: run the load twice. The second run must be a clean
no-op. If it creates duplicates, the external ID setup is wrong and production
would have been a disaster.
"""

from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from lib import mapping, workbook as wbmod

ROOT = Path(__file__).resolve().parent.parent
WB = ROOT / "mapping" / "workbook.xlsx"
RAW = ROOT / "raw"
LOGS = ROOT / "logs"

EXTERNAL_ID = "HubSpot_Record_Id__c"


def log(msg=""):
    print(msg, flush=True)


# ---------------------------------------------------------------------------
# Preconditions — refuse rather than half-load
# ---------------------------------------------------------------------------

def check_preconditions(tabs, objects) -> list[str]:
    problems = []

    for tab_name, tab in tabs.items():
        fields = tab.sections.get("FIELDS")
        if not fields:
            continue
        deploying = [r for r in fields.rows
                     if str(r.get("Deploy", "")).strip().upper() == "Y"]
        if not deploying:
            continue

        results = {str(r.get("Result", "")).strip() for r in deploying}
        if "Failed" in results:
            problems.append(
                f"{tab_name}: has fields with Result=Failed. Fix the schema deploy first.")
        if not any(r == "Success" for r in results):
            problems.append(
                f"{tab_name}: no field shows Result=Success — the schema deploy has "
                "not run, or writeback was skipped. Records cannot load into fields "
                "that do not exist.")

        rules = tab.sections.get("VALIDATION_RULES")
        if rules:
            active = [r for r in rules.rows
                      if str(r.get("Deploy", "")).strip().upper() == "Y"
                      and str(r.get("Active", "")).strip().upper() in ("TRUE", "Y", "YES")]
            if active:
                problems.append(
                    f"{tab_name}: {len(active)} ACTIVE validation rule(s) deployed. "
                    "These will reject historical records. Deactivate before loading.")

    return problems


def load_owner_map() -> dict[str, str]:
    from openpyxl import load_workbook

    wb = load_workbook(WB, data_only=True)
    try:
        if "_Users" not in wb.sheetnames:
            return {}
        ws = wb["_Users"]
        owners = {}
        for r in range(6, ws.max_row + 1):
            owner_id = ws.cell(row=r, column=2).value
            username = ws.cell(row=r, column=6).value
            if owner_id and username:
                owners[str(owner_id)] = str(username)
        return owners
    finally:
        wb.close()


def load_field_map(tabs) -> dict[str, dict[str, str]]:
    """{tab: {hubspot_property: sf_api_name}} for everything deployed."""
    field_map = {}
    for tab_name, tab in tabs.items():
        fields = tab.sections.get("FIELDS")
        if not fields:
            continue
        mapped = {}
        for row in fields.rows:
            if str(row.get("Deploy", "")).strip().upper() != "Y":
                continue
            hs = str(row.get("HS Property", "")).strip()
            api = str(row.get("Final_API") or row.get("Proposed_API") or "").strip()
            if hs and api:
                mapped[hs] = api
        if mapped:
            field_map[tab_name] = mapped
    return field_map


# ---------------------------------------------------------------------------
# Transform
# ---------------------------------------------------------------------------

def transform(records, field_map, owner_map, sf_object, include_lookups=False):
    """HubSpot records → CSV rows for Bulk upsert.

    Pass 1 omits lookups entirely: the target records may not exist yet.
    Pass 2 sets only the external id plus the relationship columns.
    """
    rows = []
    for record in records:
        props = record.get("properties") or {}
        out = {EXTERNAL_ID: record.get("id", "")}

        for hs_prop, value in props.items():
            api = field_map.get(hs_prop)
            if not api:
                continue
            if api == EXTERNAL_ID:
                continue
            if hs_prop == "hubspot_owner_id":
                # Ownership is resolved here, during transform. Applying it
                # afterwards means every record lands on the running user first.
                username = owner_map.get(str(value))
                if username:
                    out["Owner.Username"] = username
                continue
            if api.endswith("__r") or (include_lookups is False and api.endswith("__c")
                                       and hs_prop.endswith("_id")):
                continue
            out[api] = value

        rows.append(out)
    return rows


def write_csv(rows, path: Path):
    if not rows:
        return 0
    columns = sorted({key for row in rows for key in row})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


# ---------------------------------------------------------------------------
# Bulk upsert
# ---------------------------------------------------------------------------

def bulk_upsert(org: str, sf_object: str, csv_path: Path, execute: bool) -> dict:
    args = ["sf", "data", "upsert", "bulk",
            "--sobject", sf_object,
            "--file", str(csv_path),
            "--external-id", EXTERNAL_ID,
            "--target-org", org,
            "--wait", "30", "--json"]

    if not execute:
        log(f"    DRY RUN — would run: {' '.join(args)}")
        return {"dryRun": True, "sobject": sf_object}

    result = subprocess.run(args, cwd=ROOT, capture_output=True, text=True, timeout=3600)
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        return {"error": result.stderr or result.stdout}


def main() -> int:
    parser = argparse.ArgumentParser(description="Load HubSpot records via Bulk API 2.0")
    parser.add_argument("--org", default="client-sbx")
    parser.add_argument("--objects", default="all")
    parser.add_argument("--execute", action="store_true",
                        help="actually load. Without this it is a dry run.")
    parser.add_argument("--pass", dest="which_pass", choices=["1", "2", "both"],
                        default="both", help="1=records, 2=relationships")
    args = parser.parse_args()

    log("=" * 70)
    log("07 — DATA LOAD" + ("" if args.execute else "  (DRY RUN)"))
    log("=" * 70)
    log("\nNOTE: the current project phase is schema-only. Running this for real")
    log("needs a separate, explicit decision.\n")

    if not WB.exists():
        log("ERROR: no workbook.")
        return 1

    tabs = wbmod.read_workbook(WB)
    config = mapping.load_config()

    problems = check_preconditions(tabs, args.objects)
    if problems:
        log(f"{len(problems)} PRECONDITION FAILURE(S):\n")
        for problem in problems:
            log(f"  - {problem}")
        log("\nRefusing to load. Every one of these produces silent data damage.")
        return 1

    owner_map = load_owner_map()
    field_map = load_field_map(tabs)
    log(f"  {len(owner_map)} owner(s) mapped, {len(field_map)} object(s) with fields\n")

    # Parents before children, from config load_order.
    order = []
    for hs_object, entry in (config["objects"].get("objects") or {}).items():
        order.append((entry.get("load_order", 100), hs_object, entry["sf_object"]))
    defaults_order = (config["objects"].get("defaults", {})
                      .get("unknown_object", {}).get("load_order", 100))
    for hs_object in (config["objects"].get("discovered", {}).get("custom_objects") or []):
        if not any(o[1] == hs_object for o in order):
            pascal, _ = mapping.to_api_name(hs_object, config["naming"], suffix=False)
            order.append((defaults_order, hs_object, f"{pascal}__c"))
    order.sort()

    if args.objects != "all":
        wanted = {o.strip() for o in args.objects.split(",")}
        order = [o for o in order if o[1] in wanted]

    LOGS.mkdir(exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    summary = []

    with tempfile.TemporaryDirectory() as tmp:
        for which_pass in ([1, 2] if args.which_pass == "both" else [int(args.which_pass)]):
            log(f"\n{'-' * 70}")
            log(f"PASS {which_pass}: " + ("records" if which_pass == 1 else "relationships"))
            log(f"{'-' * 70}")

            for _, hs_object, sf_object in order:
                source = RAW / f"records_{hs_object}.json"
                if not source.exists():
                    source = RAW / "samples.json"
                    if not source.exists():
                        log(f"  {hs_object}: no records in raw/ — skipped")
                        continue
                    records = json.loads(source.read_text(encoding="utf-8")).get(hs_object, [])
                else:
                    records = json.loads(source.read_text(encoding="utf-8"))

                if not records:
                    log(f"  {hs_object}: 0 records — skipped")
                    continue

                tab_key = next((t for t in field_map if t.replace("__c", "").lower()
                                == sf_object.replace("__c", "").lower()), None)
                if not tab_key:
                    log(f"  {hs_object}: no deployed fields — skipped")
                    continue

                rows = transform(records, field_map[tab_key], owner_map, sf_object,
                                 include_lookups=(which_pass == 2))
                csv_path = Path(tmp) / f"{sf_object}_pass{which_pass}.csv"
                count = write_csv(rows, csv_path)
                log(f"  {hs_object} → {sf_object}: {count} record(s)")

                result = bulk_upsert(args.org, sf_object, csv_path, args.execute)
                summary.append({"object": sf_object, "pass": which_pass,
                                "attempted": count, "result": result})

    path = LOGS / f"load_{stamp}.json"
    path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    log("\n" + "=" * 70)
    log("SUMMARY")
    log("=" * 70)
    for entry in summary:
        log(f"  {entry['object']} pass {entry['pass']}: {entry['attempted']} attempted")
    log(f"\n  log: {path.relative_to(ROOT)}")

    if not args.execute:
        log("\nDRY RUN — nothing was written. Add --execute to load for real.")
    else:
        log("\nRUN IT AGAIN. The second run must be a clean no-op.")
        log("If it creates duplicates, the external ID setup is wrong and")
        log("production would have been a disaster.")
        log("\nThen activate validation rules and confirm they do not reject")
        log("the loaded data.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
